import openpyxl
from openpyxl.styles import PatternFill

from getRouteTime import getRouteTimeWithoutStops, getTimeHours, getRouteDistanceKilometers

def create_schedule_table(adresses):
    # считаем количество дней в пути
    trips = []
    totalTripTimeWithStops = 0
    for i in range(0, len(adresses)):
        if i == len(adresses) - 1:
            totalTripTimeWithStops += getTimeHours(adresses[i], adresses[0])
            trips.append(round(getTimeHours(adresses[i], adresses[0])))
        else:
            totalTripTimeWithStops += getTimeHours(adresses[i], adresses[i+1]) + 1
            trips.append(round(getTimeHours(adresses[i], adresses[i+1])))
    print("Поездки (часов) - ", trips)
    trips.reverse()

    # считаем кол-во дней
    days = round(totalTripTimeWithStops / 8)
    if days > 5:
        return 'Ошибка, количество дней превышает 5'

    # считаем расстояние, зарплату, премию, топливный сбор
    hotelPayment = 0
    dailyPayment = days * 700
    distance = getRouteDistanceKilometers(adresses)
    fuelPayment = distance / 100 * 22 * 57.25
    bonusPayment = distance * 4
    totalPayment = bonusPayment + dailyPayment

    print("Общее время поездки - ", totalTripTimeWithStops , "ч.    ")

    totalTripTimeWithOutStops = totalTripTimeWithStops - len(adresses) - 1

    travel_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    rest_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')


    # создаем новую книгу Excel
    workbook = openpyxl.Workbook()

    # выбираем активный лист
    sheet = workbook.active

    # заполняем заголовки столбцов
    sheet.cell(row=1, column=1, value="День")
    for hour in range(1, 25):
        sheet.cell(row=1, column=hour+4, value=str(hour))

    rideTimeForTheDay = 8
    rideTimeInARow = 0
    restInARow = 0
    currentTrip = trips.pop()

    # добавляем 7 строк для каждого дня недели
    days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница"]
    for i, day in enumerate(days):
        sheet.cell(row=i + 2, column=1, value=day)

        # заполняем 24 столбца для каждого дня
        for j in range(1, 25):
            if i == 0 and j <= 8:
                sheet.cell(row=i+2, column=j+4, value="Отдых").fill = rest_fill

            elif currentTrip == 0 and len(trips) == 0:
                break

            elif 4 >= currentTrip > 0 and rideTimeInARow < 4 and rideTimeForTheDay != 0:
                sheet.cell(row=i+2, column=j+4, value="Путь").fill = travel_fill
                rideTimeForTheDay -= 1
                rideTimeInARow += 1
                restInARow = 0
                currentTrip -= 1

            elif currentTrip == 0 and len(trips) > 0:
                currentTrip = trips.pop()
                sheet.cell(row=i + 2, column=j + 4, value="Отдых").fill = rest_fill
                rideTimeInARow = 0

            elif rideTimeForTheDay != 0 and rideTimeInARow < 4:
                sheet.cell(row=i+2, column=j+4, value="Путь").fill = travel_fill
                rideTimeForTheDay -= 1
                rideTimeInARow += 1
                restInARow = 0
                currentTrip -= 1

            elif rideTimeInARow >=4:
                sheet.cell(row=i+2, column=j+4, value="Отдых").fill = rest_fill
                rideTimeInARow = 0

            elif restInARow == 8:
                sheet.cell(row=i + 2, column=j + 4, value="Отдых").fill = rest_fill
                hotelPayment += 2000
                rideTimeForTheDay = 8

            elif rideTimeForTheDay == 0:
                sheet.cell(row=i + 2, column=j + 4, value="Отдых").fill = rest_fill
                restInARow += 1



    # добавляем информацию по деньгам
    sheet.cell(row=8, column=1, value="Топливный сбор")
    sheet.cell(row=9, column=1, value=fuelPayment)
    sheet.cell(row=8, column=2, value="Суточные")
    sheet.cell(row=9, column=2, value=dailyPayment)
    sheet.cell(row=8, column=3, value="Общее расстояние")
    sheet.cell(row=9, column=3, value=distance)
    sheet.cell(row=10, column=1, value="Гостиничный сбор")
    sheet.cell(row=11, column=1, value=hotelPayment)
    sheet.cell(row=10, column=2, value="Премиальные")
    sheet.cell(row=11, column=2, value=bonusPayment)
    sheet.cell(row=10, column=3, value="общая зарплата")
    sheet.cell(row=11, column=3, value=totalPayment)

    # автоматически расширяем ячейки таблицы
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length + 3

    # сохраняем таблицу
    workbook.save("schedule.xlsx")

    print("Таблица создана")
